import pandas as pd
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


file_path = "Your_File.xlsx"
log_sheet_name = "log"
param_sheet_name = "param"


log_df = pd.read_excel(file_path, sheet_name=log_sheet_name)
param_df = pd.read_excel(file_path, sheet_name=param_sheet_name)

log_df["Date If Any"] = pd.to_datetime(log_df["Date"])


repeated_rows = []


for index, row in param_df.iterrows():
    projection_value = row['Projection']+1
    log_rows_to_repeat = [log_df.iloc[i].tolist() for i in range(len(log_df))]
    repeated_rows.extend(log_rows_to_repeat * projection_value)


repeated_df = pd.DataFrame(repeated_rows, columns=log_df.columns)


trade_date_increment = relativedelta(months=1)
full_set_size = len(log_df)


month_text_mapping = {
    1: "January", 2: "February", 3: "March", 4: "April",
    5: "May", 6: "June", 7: "July", 8: "August",
    9: "September", 10: "October", 11: "November", 12: "December"
}

for index, row in repeated_df.iterrows():
    trade_date = row["Date"]  
    months_to_add = (index // full_set_size)
    new_trade_date = trade_date + relativedelta(months=months_to_add)
    repeated_df.at[index, "Date"] = new_trade_date

    
    transaction_value = repeated_df.at[index, "Transaction"]
    if isinstance(transaction_value, str):  
        for month_number, month_name in month_text_mapping.items():
            if month_name.lower() in transaction_value.lower():
                repeated_df.at[index, "Transaction"] = month_text_mapping[new_trade_date.month]
                break


repeated_df["Date"] = repeated_df["Date"].dt.strftime("%Y-%m-%d")


for index in range(len(log_df), len(repeated_df)):
    original_index = index % full_set_size
    if repeated_df.loc[index].equals(log_df.loc[original_index]):
        repeated_df.at[index, "if booked or not(ur column)"] = "Projection"
    else:
        repeated_df.at[index, "ur column"] = "Projected"


output_file_path = f"projected_for_{projection_value-1}_month.xlsx"
repeated_df.to_excel(output_file_path, index=False)


workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'log'


for row in dataframe_to_rows(repeated_df, index=False, header=True):
    worksheet.append(row)


for column_cells in worksheet.columns:
    for cell in column_cells:
        cell.alignment = Alignment(horizontal="right")


for column in worksheet.columns:
    max_length = max(len(str(cell.value)) for cell in column)
    adjusted_width = (max_length + 2) * 1.2
    worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

# Save the customized workbook
workbook.save(output_file_path)
